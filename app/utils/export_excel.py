from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import os

def export_products_to_excel(products, logs, filename="products_export.xlsx"):
    """
    Xuất danh sách sản phẩm và log ra file Excel
    
    Args:
        products: Danh sách sản phẩm
        logs: Danh sách log thay đổi
        filename: Tên file Excel
    """
    wb = Workbook()
    
    # Sheet 1: Danh sách sản phẩm
    ws_products = wb.active
    ws_products.title = "Sản phẩm"
    
    # Header cho sheet sản phẩm
    headers = ["ID", "Tên sản phẩm", "SKU", "Giá", "Số lượng", "Danh mục", "Mô tả", "Ảnh", "Ngày tạo", "Ngày cập nhật"]
    for col, header in enumerate(headers, 1):
        cell = ws_products.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Dữ liệu sản phẩm
    for row, product in enumerate(products, 2):
        ws_products.cell(row=row, column=1, value=product.id)
        ws_products.cell(row=row, column=2, value=product.name)
        ws_products.cell(row=row, column=3, value=product.sku)
        ws_products.cell(row=row, column=4, value=product.price)
        ws_products.cell(row=row, column=5, value=product.quantity)
        ws_products.cell(row=row, column=6, value=product.category)
        ws_products.cell(row=row, column=7, value=product.description)
        ws_products.cell(row=row, column=8, value=", ".join(product.images) if product.images else "")
        ws_products.cell(row=row, column=9, value=product.created_at.strftime("%Y-%m-%d %H:%M:%S") if product.created_at else "")
        ws_products.cell(row=row, column=10, value=product.updated_at.strftime("%Y-%m-%d %H:%M:%S") if product.updated_at else "")
    
    # Sheet 2: Lịch sử thay đổi
    ws_logs = wb.create_sheet("Lịch sử thay đổi")
    
    # Header cho sheet log
    log_headers = ["ID", "ID Sản phẩm", "Hành động", "Trường thay đổi", "Giá trị cũ", "Giá trị mới", "Người thay đổi", "Thời gian"]
    for col, header in enumerate(log_headers, 1):
        cell = ws_logs.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Dữ liệu log
    for row, log in enumerate(logs, 2):
        ws_logs.cell(row=row, column=1, value=log.id)
        ws_logs.cell(row=row, column=2, value=log.product_id)
        ws_logs.cell(row=row, column=3, value=log.action)
        ws_logs.cell(row=row, column=4, value=log.field_name)
        ws_logs.cell(row=row, column=5, value=log.old_value)
        ws_logs.cell(row=row, column=6, value=log.new_value)
        ws_logs.cell(row=row, column=7, value=log.changed_by)
        ws_logs.cell(row=row, column=8, value=log.created_at.strftime("%Y-%m-%d %H:%M:%S") if log.created_at else "")
    
    # Điều chỉnh độ rộng cột
    for ws in [ws_products, ws_logs]:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Lưu file
    export_dir = "exports"
    os.makedirs(export_dir, exist_ok=True)
    filepath = os.path.join(export_dir, filename)
    wb.save(filepath)
    
    return filepath 